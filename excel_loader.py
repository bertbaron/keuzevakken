from openpyxl.reader.excel import load_workbook

from model import Data, HandledException, Config, ClassConfig, Course, Student, ResultRecord


class ExcelLoaderException(HandledException):
    pass


_META_SHEETS = ["config", "vakken"]


class ExcelLoader:
    def __init__(self, inputpath: str, previous: str | None):
        self.input = inputpath
        self.previous = previous

    def load(self) -> Data:
        data = Data()
        wb = load_workbook(self.input, data_only=True)

        self._parse_config(data, wb)
        self._parse_courses(data, wb)
        self._parse_classes(data, wb)

        if self.previous:
            wb = load_workbook(self.previous, data_only=True)
            self._parse_previous(data, wb)

        return data

    def _parse_config(self, data: Data, wb):
        ws = self._get_sheet(wb, "Config")
        periods = int(ws.cell(row=(self._find_in_column(ws, 1, "Periodes")), column=2).value)
        classes = [ClassConfig(name, name) for name in wb.sheetnames
                   if name.lower() not in _META_SHEETS and not name.startswith("_")]

        def parse_pairs(ws, value):
            r, c = self._find_in_sheet(ws, value)
            r += 1
            pairs = []
            while ws.cell(row=r, column=c).value:
                pairs.append([ws.cell(row=r, column=c).value, ws.cell(row=r, column=c + 1).value])
                r += 1
            return pairs

        together = parse_pairs(ws, "Samen")
        apart = parse_pairs(ws, "Apart")

        data.config = Config(periods, classes, together, apart)

    def _parse_courses(self, data: Data, wb):
        ws = self._get_sheet(wb, "Vakken")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            code = row[0].value
            if not code:
                break
            size = int(row[1].value)
            availability = "".join(["1" if cell.value == "x" else "0" for cell in row[2:data.config.periods + 2]])
            data.courses.append(Course(code, size, availability))

    def _parse_classes(self, data: Data, wb):
        for cl in data.config.classes:
            ws = self._get_sheet(wb, cl.code)
            self._parse_class(data, ws)

    def _parse_class(self, data: Data, ws):
        students = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            name = row[0].value
            if not name:
                break
            choices = [cell.value for cell in row[1:] if cell.value]
            students.append(Student(name, choices))
        data.add_students(ws.title, students)

    def _parse_previous(self, data: Data, wb):
        previous_records = []
        for name in wb.sheetnames:
            if name.lower() in _META_SHEETS or name.startswith("_"):
                continue
            ws = self._get_sheet(wb, name)

            row = self._find_in_column(ws, 1, "Naam") + 1
            while ws.cell(row=row, column=1).value:
                name = ws.cell(row=row, column=1).value
                courses = []
                for col in range(2, data.config.periods + 2):
                    courses.append(ws.cell(row=row, column=col).value)
                    previous_records.append(ResultRecord(name, courses, 0))
                row += 1
        data.previous_result = previous_records

    def _get_sheet(self, wb, sheet_name):
        if sheet_name not in wb.sheetnames:
            raise ExcelLoaderException(f"Werkblad '{sheet_name}' niet gevonden")
        return wb[sheet_name]

    def _find_in_sheet(self, ws, value) -> tuple[int, int]:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).lower() == value.lower():
                    return cell.row, cell.column
        raise ExcelLoaderException(f"'{value}' niet gevonden in werkblad {ws.title}")

    def _find_in_column(self, ws, col, value) -> int:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
            if row[0].value and row[0].value.lower() == value.lower():
                return row[0].row
        raise ExcelLoaderException(f"'{value}' niet gevonden in werkblad {ws.title}")

    # def _find_value(sheet, value) -> tuple[int, int]:
    #     for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    #         if row[0].value and row[0].value.lower() == value:
    #             return row[0].row, 2
    #     raise HandledException(f"Value '{value}' not found in sheet {sheet.title}")
