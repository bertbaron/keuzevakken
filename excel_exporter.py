import colorsys

from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook

from model import Data, ClassConfig
from util import make_backup

CELL_WIDTH_STUDENT_NAME = 25
CELL_WIDTH_COURSE_NAME = 16
CELL_WIDTH_BREAK_COLUMN = 2

HUE_STEP = 24
COLOR_VARIANTS = (
    # (lightness, saturation)
    (80, 55),
    (75, 70),
)

class ExcelExporter:
    def export(self, data: Data, filename: str):
        wb = Workbook()
        wb.remove(wb.active)

        # classes = set(data.student_to_class.values())
        for class_config in data.config.classes:
            # class_config = next(cl for cl in data.config.classes if cl.code == code)
            self._add_class_sheet(wb, data, class_config)

        self._add_courses_sheet(wb, data)

        make_backup(filename)
        wb.save(filename)

    def _add_class_sheet(self, wb: Workbook, data: Data, cl: ClassConfig):
        ws = wb.create_sheet(title=cl.name)

        ws.append([cl.name, "keuzevakken indeling"])
        ws.cell(row=1, column=1).font = ws.cell(row=1, column=1).font.copy(size=14)

        ws.append([])
        ws.append(["Naam", "Periode 1", "Periode 2", "Periode 3", "Periode 4", "", "Reserve"])

        code_to_index = {course.code: i for i, course in enumerate(data.courses)}

        for record in data.enriched_result:
            if data.student_to_class[record.student] == cl.code:
                courses = record.assigned[:data.config.periods] + [None] + record.assigned[data.config.periods:]
                row = [record.student]
                for course in courses:
                    if course:
                        row.append(data.get_course_name(course.code) + ("*" if course.original_reserve else ""))
                    else:
                        row.append("")
                ws.append(row)

                # color the cells based on the course
                for i, course in enumerate(courses):
                    if course:
                        if course.fits:
                            fill = self._solid_fill(self._color_for_index(code_to_index[course.code]))
                        else:
                            # intense red
                            fill = self._solid_fill(self._hls_color(0, 60, 100))
                        ws.cell(row=ws.max_row, column=i + 2).fill = fill

        ws.append([])
        ws.append(["* = oorspronkelijke reservekeuze"])

        ws.column_dimensions["A"].width = CELL_WIDTH_STUDENT_NAME
        for i in range(1, 8):
            width = CELL_WIDTH_BREAK_COLUMN if i == data.config.periods + 1 else CELL_WIDTH_COURSE_NAME
            ws.column_dimensions[chr(65 + i)].width = width

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.print_area = ws.dimensions

    def _add_courses_sheet(self, wb: Workbook, data: Data):
        ws = wb.create_sheet(title="Vakken")

        ws.append(["Vak", "Grootte"] + [f"Periode {i + 1}" for i in range(data.config.periods)])

        for y, course in enumerate(data.courses):
            record = [course.name, course.size]
            for p in range(data.config.periods):
                count = 0
                for rr in data.enriched_result:
                    if rr.assigned[p] and rr.assigned[p].code == course.code:
                        count += 1
                if count > 0 or course.availability[p]:
                    record.append(count)
                else:
                    record.append("")
            ws.append(record)

        for i, course in enumerate(data.courses):
            y = i + 2
            for x in range(3, data.config.periods + 3):
                cell = ws.cell(row=y, column=x)
                count = int(cell.value) if cell.value else 0
                size = course.size if course.availability[x - 3] else 0
                if count > size:
                    cell.fill = self._solid_fill(self._hls_color(0, 60, 100))
                elif count > 0:
                    ratio = count / size
                    cell.fill = self._solid_fill(self._hls_color(0, 100 - ratio * 100 / 2, 0))

        ws.column_dimensions["A"].width = CELL_WIDTH_COURSE_NAME
        for p in range(data.config.periods + 1):
            ws.column_dimensions[chr(66 + p)].width = 10

        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.print_area = ws.dimensions

    def _solid_fill(self, color: str):
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _color_for_index(self, index: int):
        hue = index * HUE_STEP % 360
        variant = index * HUE_STEP // 360
        return self._friendly_background(hue, variant)

    def _color_for_hash(self, code: str):
        """
        Generate a color based on the hash code of the course code. This results in the same color for the same course
        in all sheets.
        """
        hashcode = 0
        for c in code:
            hashcode = (31 * hashcode + ord(c)) & 0xFFFFFFFF

        hue = hashcode % 360
        return self._friendly_background(hue)

    def _friendly_background(self, hue: int, variant=0):
        lightness, saturation = COLOR_VARIANTS[(variant % len(COLOR_VARIANTS))]
        return self._hls_color(hue, lightness, saturation)

    def _hls_color(self, hue: int, lightness, saturation):
        """
        :param hue: 0-360 degrees
        :param lightness: 0-100%
        :param saturation: 0-100%
        :return:
        """
        rgb = colorsys.hls_to_rgb(hue / 360, lightness / 100, saturation / 100)
        return f"{int(rgb[0] * 255):02X}{int(rgb[1] * 255):02X}{int(rgb[2] * 255):02X}"
